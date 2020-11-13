from tkinter import *
import pandas as pd
from openpyxl import load_workbook
from datetime import date, timedelta
from os.path import isfile
from os import system
from shutil import copy
from morning_report import generate_morning_report

global events_list, event
customers_dict = {'CVG': ['ABX', 'ATI AMZ', 'ATI DHL', 'DHL', '21 Air', 'Frontier', 'Cargojet', 'Aerologic', 'Commutair', 'Air Georgian', 'MESA', 'Southwest', 'Republic', 'United Airlines', 'National', 'Northern Air Cargo', 'Swift', 'Sky West'], 'ILN': ['ABX', 'ATI', 'Atlas', 'Sun Country'], 'MIA': ['ABX', 'ATI', 'DHL', 'Cargojet', 'Northern Air Cargo', 'Amerijet', 'Sunwing']}
loc_events_dict = {'CVG': ['Turns', 'Borescopes', 'Meet and Greet'], 'ILN': ['Turns', 'Borescopes', 'Weekly Parts Runner Per Turn', 'Weekly Comat per Turn'], 'MIA': ['Turns', 'Borescopes', 'Labor OT']}
events_dict = {'ABX': ['Turns', 'Borescopes', 'Labor OT', 'Weekly Comat per turn'], 'ATI AMZ': ['Turns', 'Labor OT'], 'ATI DHL': ['Turns', 'Borescopes', 'Labor OT'], 'ATI': ['Turns', 'Labor OT', 'Weekly Parts Runner per turn'], 'DHL': ['Turns', 'Borescopes'], '21 Air': ['Turns', 'Service Check', 'Labor OT'], 'Atlas': ['Turns'], 'Frontier': ['Meet and Greet', 'Periodic Check', 'Wakeup Check', 'Service Check', 'Heavy Check', 'Daily Check', 'Cold Weather Check', 'O2 Charge'], 'Cargojet': ['Turns', 'Service Check', 'Pre-Departure', 'Transit Check', 'Weekly Check'], 'Aerologic': ['ETOPS Check', 'Daily Check', 'Weekly Check'], 'Commutair': ['On Call Service', 'Labor OT'], 'Air Georgian': ['On Call Service', 'Labor OT'], 'MESA': ['On Call Service', 'Labor OT'], 'Southwest': ['On Call Service', 'Labor OT'], 'Republic': ['On Call Service', 'Labor OT'], 'United Airlines': ['On Call Service', 'Labor OT'], 'National': ['Turns'], 'Northern Air Cargo': ['Turns'], 'Swift': ['Meet and Greet', 'Service Check'], 'Amerijet': ['Turns', 'Labor OT]'], 'Sunwing': ['Turns', 'Labor OT'], 'Sky West': ['On Call Service', 'Labor OT'], 'Sun Country': ['Turns']}


def add_event(win, loc, service, tail_num):
    # print(service, tail_num, company, customer)
    df = pd.read_excel(path, sheet_name='Events', index_col=[0])
    tail_line = [tail_num, customer.get(), service]
    wb = load_workbook(path)
    tail_ws = wb['Tail Numbers']
    # tail_list = []
    # for tail in tail_ws['A']:
    #     tail_list.append(tail.value)
    # if not (tail_num in tail_list):
    tail_ws.append(tail_line)
    wb.save(path)
    df.loc[customer.get(), service] += 1
    with pd.ExcelWriter(path, engine="openpyxl", mode='a') as writer:
        workBook = writer.book
        try:
            workBook.remove(workBook['Events'])
        except KeyError:
            print("worksheet doesn't exist")
        df.to_excel(writer, sheet_name='Events')
        writer.save()
    generate_morning_report(report_date.get())
    win.destroy()
    set_event(loc)


def add_training(win, loc, employee_num, length, description):
    line = [customer.get(), int(employee_num), float(length), description]
    # print(line)
    wb = load_workbook(path)
    ws = wb['Training']
    ws.append(line)
    wb.save(path)
    generate_morning_report(report_date.get())
    win.destroy()
    set_training(loc)


def add_delay(win, loc, tail_num, issue, description, notes):
    line = [customer.get(), tail_num, issue, description, notes]
    # print(line)
    wb = load_workbook(path)
    ws = wb['Delays Cancellations']
    ws.append(line)
    wb.save(path)
    generate_morning_report(report_date.get())
    win.destroy()
    set_delay(loc)


def add_aircraft_in_work(win, loc, tail_num, work_order, notes):
    line = [tail_num, customer.get(), int(work_order), notes]
    # print(line)
    wb = load_workbook(path)
    ws = wb['Aircraft in Work']
    ws.append(line)
    wb.save(path)
    generate_morning_report(report_date.get())
    win.destroy()
    set_aircraft_in_work(loc)


def add_ee_status(win, loc, status, count, notes):
    df = pd.read_excel(path, sheet_name='EE Status', index_col=[0])
    df.loc[status, '#'] += count
    df.loc[status, 'Notes'] = notes
    with pd.ExcelWriter(path, engine="openpyxl", mode='a') as writer:
        workBook = writer.book
        try:
            workBook.remove(workBook['EE Status'])
        except KeyError:
            print("worksheet doesn't exist")
        df.to_excel(writer, sheet_name='EE Status')
        writer.save()
    generate_morning_report(report_date.get())
    win.destroy()
    set_ee_status(loc)


def open_excel():
    make_file()
    system('start "excel" "{}"'.format(path))


def set_event(loc):
    global events_list, event
    make_file()
    customers = customers_dict[loc]
    if not customer.get() or not customer.get() in customers:
        customer.set(customers[0])
    window = Toplevel(root, bg='White')
    window.title('Add Service')
    loc_label = Label(window, text='Location: ', bg='White').grid(row=0)
    loc_label = Label(window, text=loc, bg='White').grid(row=0, column=1)
    tail_num_label = Label(window, text='Tail Number: ', bg='White').grid(row=1)
    tail_num_entry = Entry(window)
    tail_num_entry.grid(row=1, column=1)
    customer_label = Label(window, text='Customer: ', bg='White').grid(row=2)
    customer_list = OptionMenu(window, customer, *customers)
    customer_list.grid(row=2, column=1)
    events_label = Label(window, text='Service: ', bg='White').grid(row=3)
    event = StringVar()
    event.set(events_dict[customer.get()][0])
    events_list = OptionMenu(window, event, *events_dict[customer.get()])
    events_list.grid(row=3, column=1)
    # num_to_add_label = Label(window, text='Number to add: ', bg='White').grid(row=4)
    # num_to_add_entry = Entry(window)
    # num_to_add_entry.grid(row=4, column=1)
    customer.trace('w', lambda a, b, c: update_events())
    add_button = Button(window, text='Add', command=lambda: add_event(window, loc, event.get(), tail_num_entry.get())).grid(row=4)  # num_to_add_entry.get())).grid(row=5)
    cancel_button = Button(window, text='Cancel', command=window.destroy).grid(row=4, column=1)


def set_training(loc):
    make_file()
    customers = customers_dict[loc]
    window = Toplevel(root, bg='White')
    window.title('Add Training')
    loc_label = Label(window, text='Location: ', bg='White').grid(row=0)
    loc_label = Label(window, text=loc, bg='White').grid(row=0, column=1)
    customer_label = Label(window, text='Company: ', bg='White').grid(row=1)
    # customer = StringVar()
    # customer.set(customers[0])
    customer_list = OptionMenu(window, customer, *customers)
    customer_list.grid(row=1, column=1)
    employee_num_label = Label(window, text='Employee Number: ', bg='White').grid(row=2)
    employee_num_entry = Entry(window)
    employee_num_entry.grid(row=2, column=1)
    length_label = Label(window, text='Length (hours): ', bg='White').grid(row=3)
    length_entry = Entry(window)
    length_entry.grid(row=3, column=1)
    description_label = Label(window, text='Description: ', bg='White').grid(row=4)
    description_entry = Entry(window)
    description_entry.grid(row=4, column=1)
    add_button = Button(window, text='Add',
                        command=lambda: add_training(window, loc, employee_num_entry.get(), length_entry.get(),
                                                     description_entry.get())).grid(row=5)
    cancel_button = Button(window, text='Cancel', command=window.destroy).grid(row=5, column=1)


def set_delay(loc):
    customers = customers_dict[loc]
    make_file()
    window = Toplevel(root, bg='White')
    window.title('Add Delay')
    loc_label = Label(window, text='Location: ', bg='White').grid(row=0)
    loc_label = Label(window, text=loc, bg='White').grid(row=0, column=1)
    tail_num_label = Label(window, text='Tail Number: ', bg='White').grid(row=1)
    tail_num_entry = Entry(window)
    tail_num_entry.grid(row=1, column=1)
    customer_label = Label(window, text='Customer: ', bg='White').grid(row=2)
    # customer = StringVar()
    # customer.set(customers[0])
    customer_list = OptionMenu(window, customer, *customers)
    customer_list.grid(row=2, column=1)
    # length_label = Label(window, text='Length (hours): ', bg='White').grid(row=1)
    # length_entry = Entry(window).grid(row=1, column=1)
    issue_label = Label(window, text='Issue: ', bg='White').grid(row=3)
    issue_entry = Entry(window)
    issue_entry.grid(row=3, column=1)
    description_label = Label(window, text='Preventable? : ', bg='White').grid(row=4)
    description = StringVar()
    description.set('Preventable')
    description_list = OptionMenu(window, description, 'Preventable', 'Excusable', 'Non-Excusable')
    description_list.grid(row=4, column=1)
    notes_label = Label(window, text='Notes: ', bg='White').grid(row=5)
    notes_entry = Entry(window)
    notes_entry.grid(row=5, column=1)
    add_button = Button(window, text='Add',
                        command=lambda: add_delay(window, loc, tail_num_entry.get(), issue_entry.get(),
                                                  description.get(), notes_entry.get())).grid(row=6)
    cancel_button = Button(window, text='Cancel', command=window.destroy).grid(row=6, column=1)


def set_aircraft_in_work(loc):
    make_file()
    customers = customers_dict[loc]
    window = Toplevel(root, bg='White')
    window.title('Add Aircraft in Work')
    loc_label = Label(window, text='Location: ', bg='White').grid(row=0)
    loc_label = Label(window, text=loc, bg='White').grid(row=0, column=1)
    tail_num_label = Label(window, text='Tail Number: ', bg='White').grid(row=1)
    tail_num_entry = Entry(window)
    tail_num_entry.grid(row=1, column=1)
    customer_label = Label(window, text='Customer: ', bg='White').grid(row=2)
    # customer = StringVar()
    # customer.set(customers[0])
    customer_list = OptionMenu(window, customer, *customers)
    customer_list.grid(row=2, column=1)
    # dm_di_label = Label(window, text='DM/DI: ', bg='White').grid(row=2)
    # dm_di_entry = Entry(window)
    # dm_di_entry.grid(row=2, column=1)
    work_order_label = Label(window, text='Work Order: ', bg='White').grid(row=3)
    work_order_entry = Entry(window)
    work_order_entry.grid(row=3, column=1)
    notes_label = Label(window, text='Notes: ', bg='White').grid(row=4)
    notes_entry = Entry(window)
    notes_entry.grid(row=4, column=1)
    add_button = Button(window, text='Add',
                        command=lambda: add_aircraft_in_work(window, loc, tail_num_entry.get(), work_order_entry.get(), notes_entry.get())).grid(row=5)
    cancel_button = Button(window, text='Cancel', command=window.destroy).grid(row=5, column=1)


def set_ee_status(loc):
    make_file()
    window = Toplevel(root, bg='White')
    window.title('Add EE Status')
    loc_label = Label(window, text='Location: ', bg='White').grid(row=0)
    loc_label = Label(window, text=loc, bg='White').grid(row=0, column=1)
    status_label = Label(window, text='Status: ', bg='White').grid(row=1)
    status = StringVar()
    status.set('Injuries/Damages')
    status_list = OptionMenu(window, status, 'Injuries/Damages', 'FAA Visits', 'Travel Assignments', 'Employees out of service', 'Delay')
    status_list.grid(row=1, column=1)
    count_label = Label(window, text='Count: ', bg='White').grid(row=2)
    count_entry = Entry(window)
    count_entry.grid(row=2, column=1)
    notes_label = Label(window, text='Notes: ', bg='White').grid(row=3)
    notes_entry = Entry(window)
    notes_entry.grid(row=3, column=1)
    add_button = Button(window, text='Add', command=lambda: add_ee_status(window, loc, status.get(), count_entry.get(), notes_entry.get())).grid(row=4)  # num_to_add_entry.get())).grid(row=5)
    cancel_button = Button(window, text='Cancel', command=window.destroy).grid(row=4, column=1)


def update_vars(*args):
    global path, path_start
    d = report_date.get()
    loc = location.get()
    # customer.set(customers_dict[loc][0])
    updated_path_start = r'G:\Line Reports\Reports\{} Daily Events'.format(loc)
    p = updated_path_start + r'\{} {} Line Maintenance Report.xlsx'.format(loc, d)
    path_start = updated_path_start
    path = p
    print(path)
    return updated_path_start, p


def update_events():
    global event, events_list
    if events_list and events_list['menu']:
        events = events_dict[customer.get()]
        # print(events)
        # print(events_list)
        event.set(events[0])
        events_list['menu'].delete(0, 'end')
        for e in events:
            events_list['menu'].add_command(label=e, command=lambda ev=e: event.set(ev))


def make_file():
    if not (isfile(path)):
        copy('Line Maintenance Report.xlsx', path)


root = Tk()
root.title('Line Maintenance Report Form')
customer = StringVar()

report_date = StringVar()
report_date.set(date.today() - timedelta(days=1))
report_date.trace('w', update_vars)

location = StringVar()
location.trace('w', update_vars)
location.set('CVG')
location_list = OptionMenu(root, location, *customers_dict.keys())
location_list.grid(row=0, column=0)

path_start, path = update_vars()

add_event_button = Button(root, text='Add Event', command=lambda: set_event(location.get())).grid(row=0, column=1)
add_training_button = Button(root, text='Add Training', command=lambda: set_training(location.get())).grid(row=0, column=2)
add_delay_button = Button(root, text='Add Delay', command=lambda: set_delay(location.get())).grid(row=0, column=3)
add_aircraft_button = Button(root, text='Add Aircraft in Work', command=lambda: set_aircraft_in_work(location.get())).grid(row=0, column=4)
add_ee_status_button = Button(root, text='Add EE Status', command=lambda: set_ee_status(location.get())).grid(row=0, column=5)
open_excel_button = Button(root, text='Open Excel File', command=open_excel).grid(row=0, column=6)
report_date_label = Label(root, text='Report Date: ').grid(row=0, column=7)
report_date_entry = Entry(root, textvariable=report_date)
report_date_entry.grid(row=0, column=8)
generate_morning_report(date.today() - timedelta(days=1))

root.mainloop()
